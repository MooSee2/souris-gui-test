import functools
import glob
from datetime import datetime as dt
from pathlib import Path
from typing import Union

import numpy as np
import pandas as pd
from loguru import logger
from openpyxl import load_workbook

import souris.core.configs as cfg


def get_newest_log(path: Path) -> Path:
    """Finds most recent file in directory.

    Parameters
    ----------
    path : Path
        pathlib.Path filepath to directory to search

    Returns
    -------
    Path
        pathlib.Path of most recent file in directory
    """
    import os

    logger.debug(f"CWD: {os.getcwd()}")
    files = list(path.iterdir())
    return max(files, key=lambda file: file.stat().st_mtime)


def logger_wraps(*, entry=True, exit=True, level="DEBUG"):
    """Function decorator to log function information

    Parameters
    ----------
    entry : bool, optional
        _description_, by default True
    exit : bool, optional
        _description_, by default True
    level : str, optional
        _description_, by default "DEBUG"
    """

    def wrapper(func):
        name = func.__name__

        @functools.wraps(func)
        def wrapped(*args, **kwargs):
            logger_ = logger.opt(depth=1)
            if entry:
                logger_.log(level, "Entering '{}' (args={}, kwargs={})", name, args, kwargs)
            result = func(*args, **kwargs)
            if exit:
                logger_.log(level, "Exiting '{}' (result={})", name, result)
            return result

        return wrapped

    return wrapper


def rename_monthly_index(data: Union[pd.Series, pd.DataFrame]) -> Union[pd.Series, pd.DataFrame]:
    """Calculate monthly values of a datetime indexed pandas Series

    Parameters
    ----------
    dataframe : pd.Series
        Series to calculate monthly values with

    Returns
    -------
    pd.Series
        Series of monthly values.  Index returned in the format yyyy-mm. e.g. January = 2022-1, February = 2022-2...
    """
    if isinstance(data, pd.DataFrame):
        return data.set_index(data.index.strftime("%Y-%m"))  # type: ignore
    if isinstance(data, pd.Series):
        data.index = data.index.strftime("%Y-%m")  # type: ignore
        return data


def check_nan_gaps(
    data: pd.DataFrame,
    staid: str,
) -> pd.DatetimeIndex:
    """Checks for nan values in first column and returns DateTimeIndex of missing values.

    Parameters
    ----------
    data : pd.DataFrame
        DateTimeIndex'ed DataFrame

    Returns
    -------
    DateTimeIndex
        DateTimeIndex of missing values
    """
    mask = data.iloc[:, 0].isna()
    return pd.DataFrame({f"{staid}_gap_dates": data[mask].index})


def issue_gap_warning(data: dict, freq: str) -> None:
    for station, gaps in data.items():
        if not gaps.empty:
            logger.warning(f"Warning: Data gaps found at {station}  Freq: {freq}")
    return


def join_data(data: dict) -> pd.DataFrame:
    joined_df = pd.DataFrame()
    for item in data:
        joined_df.join(item)
    return joined_df


def penman(
    dataframe: pd.DataFrame,
    wind: str,
    temp: str,
    rel_hum: str,
    rad: str,
    ELEV: int,
) -> pd.DataFrame:
    """Returns the Penman calculation.

    Parameters
    ----------
    wind : pd.Series
        Units: m/s
    temp : pd.Series
        Units: C
    rel_hum : pd.Series
        Units: W/m2
    rad : pd.Series
        Units: MJ/m2
    ELEV : int
        Units: meters

    Returns
    -------
    pd.DataFrame

    Notes
    -----

    """
    ANNEMOMETER_HEIGHT = 3
    dataframe = dataframe.copy()

    # wind speed [km/h] from wind speed [m/s]
    dataframe["daily_wind"] = dataframe[wind] * 3.6

    # latent heat of vaporization of water [cal/g] from temperature [C]
    dataframe["latent_heat"] = 597.3 - 0.566 * dataframe[temp]

    # saturated vapor pressure [mb] from temperature [C]
    dataframe["sat_vapor"] = 6.11 * np.exp((17.26 * dataframe[temp]) / (dataframe[temp] + 237.3))

    # radiation [MJ/m2] from radiation [W/m2]
    dataframe["daily_rad"] = dataframe[rad] * 0.0864

    # pressure [mb] from elevation [m]
    dataframe["atm_pressure"] = 1013.25 * (1 - 0.0000225577 * ELEV) ** 5.25588

    # dewpoint temperautre [C] from relative humidity and temperature [C]
    dataframe["daily_dew"] = 237.3 / (17.27 / (np.log(dataframe[rel_hum] / 100) + 17.27 * dataframe[temp] / (237.3 + dataframe[temp])) - 1)

    # dewpoint vapor pressure [mb] from temperature [C]
    dataframe["vap_pressure"] = 6.11 * (np.exp((17.26 * dataframe["daily_dew"]) / (dataframe["daily_dew"] + 237.3)))

    # wind corrected to a height of 7.6 m above the ground [km/h] from wind [km/h] and
    # annemometer height [m] (set to 3 m as that's ECCC and USGS standard)
    dataframe["adj_wind"] = dataframe["daily_wind"] * (7.62 / ANNEMOMETER_HEIGHT) ** 0.25

    # rate of change of vapor pressure at temp [mb/C] from latent hear of vaporization [cal/g],
    # saturated vapor pressure [mb], and temp [C]
    dataframe["vapor_change"] = 18.02 * dataframe["latent_heat"] * dataframe["sat_vapor"] / 1.9872 / (273.16 + dataframe[temp]) ** 2

    # net radiation in units of evap [mm] from radiation [MJ/m2] and latent heat of vaporization of water [cal/g]
    dataframe["net_rad"] = 238.95 * dataframe["daily_rad"] / dataframe["latent_heat"]

    # psychrometric constant [mb/C] from pressure [mb] and latent heat of vaporization of water [cal/g]
    dataframe["gamma"] = 0.24 * dataframe["atm_pressure"] / 0.622 / dataframe["latent_heat"]

    # aerodynamic estimate of evaporation [mm] from saturated vapor pressure [mb], dewpoint vapor pressure [mb],
    # corrected wind [km/h], and elevation [m]
    dataframe["Ea"] = 10.1 / 30.4 * (dataframe["sat_vapor"] - dataframe["vap_pressure"]) * (1 + 0.062139 * dataframe["adj_wind"]) * (1 + 0.0000328084 * ELEV)

    # Penman evaporation [mm] from vapor pressure change at temp [mb/C], net radiation in evap [mm], psychrometric
    # constant [mb/C], and aerodynamic evaporation [mm]
    dataframe["penman"] = (dataframe["vapor_change"] * dataframe["net_rad"] + dataframe["gamma"] * dataframe["Ea"]) / (dataframe["vapor_change"] + dataframe["gamma"])
    return dataframe[["penman"]]


def souris_excel_writer(
    reported_flows: cfg.Boxes,
    dates: cfg.Dates,
    boxes: dict,
    daily_data: pd.DataFrame,
    monthly_elev_data: pd.DataFrame,
    daily_meteo_data: dict,
    approval_dict: dict,
    override_data: Union[None, pd.DataFrame],
    report_template: Union[str, Path],  # BLANK Souris Summary Report Template
):
    if isinstance(report_template, str):
        report_template = Path(report_template)
    # if month or day number is a single digit, add a zero to the front of the string
    app_range = f"for {dates.start_pull} to {dates.end_pull}"
    dt_now = dt.now().strftime("%Y-%m-%d %H.%M.%S.%f")
    filename = Path(f"reports/Souris Natural Flow Report {app_range} on {dt_now}.xlsx")

    ts_data_sheet_name = "Input TS data"
    original_ts_sheet_name = "Override Data"

    wb = load_workbook(report_template)
    summary_sheet = wb["Souris Natural Flow Summary"]
    config_sheet = wb["toml Settings"]
    reported_flows_sheet = wb["Input Reported Flows"]
    log_sheet = wb["Log"]
    ts_data_sheet = wb[ts_data_sheet_name]
    original_ts_data_sheet = wb[original_ts_sheet_name]
    

    # Header
    summary_sheet["B5"] = f"For the period of {app_range}"  # type: ignore erroneous pylance subscript error

    # --------------------------------------------------------------------------------------------------#
    #                                       Boxes                                                       #
    # --------------------------------------------------------------------------------------------------#

    # Larson Reservoir
    summary_sheet["B15"] = boxes["box1"]  # type: ignore erroneous pylance subscript error
    summary_sheet["C15"] = boxes["box2"]  # type: ignore erroneous pylance subscript error
    summary_sheet["D15"] = boxes["box3"]  # type: ignore erroneous pylance subscript error
    # summary_sheet["E15"] = boxes["box4"]  # type: ignore erroneous pylance subscript error
    # reported_flows_sheet["C10"] = boxes["box4"]  # type: ignore erroneous pylance subscript error

    # Boundary Reservoir
    summary_sheet["F14"] = boxes["box5a"]  # type: ignore erroneous pylance subscript error
    summary_sheet["F15"] = 0  # type: ignore erroneous pylance subscript error
    # reported_flows_sheet["D10"] = boxes["box5a"]  # type: ignore erroneous pylance subscript error
    reported_flows_sheet["D10"] = 0  # type: ignore erroneous pylance subscript error
    summary_sheet["G15"] = boxes["box6"]  # type: ignore erroneous pylance subscript error

    # summary_sheet["H15"] = boxes["box7"]  # type: ignore erroneous pylance subscript error
    # reported_flows_sheet["E10"] = boxes["box7"]  # type: ignore erroneous pylance subscript error
    summary_sheet["I15"] = boxes["box8"]  # type: ignore erroneous pylance subscript error
    summary_sheet["J15"] = boxes["box9"]  # type: ignore erroneous pylance subscript error
    summary_sheet["K15"] = boxes["box10"]  # type: ignore erroneous pylance subscript error
    summary_sheet["L15"] = reported_flows.box_11  # type: ignore erroneous pylance subscript error
    reported_flows_sheet["F10"] = reported_flows.box_11  # type: ignore erroneous pylance subscript error
    summary_sheet["M15"] = reported_flows.box_12  # type: ignore erroneous pylance subscript error
    reported_flows_sheet["G10"] = reported_flows.box_12  # type: ignore erroneous pylance subscript error
    summary_sheet["N15"] = boxes["box13"]  # type: ignore erroneous pylance subscript error

    # Nickle Lake Reservoir
    summary_sheet["B26"] = boxes["box14"]  # type: ignore erroneous pylance subscript error
    summary_sheet["C26"] = boxes["box15"]  # type: ignore erroneous pylance subscript error
    summary_sheet["D26"] = reported_flows.box_16  # type: ignore erroneous pylance subscript error
    reported_flows_sheet["H10"] = reported_flows.box_16  # type: ignore erroneous pylance subscript error
    summary_sheet["E26"] = boxes["box17"]  # type: ignore erroneous pylance subscript error

    # City of Wayburn Return Flow
    summary_sheet["F26"] = reported_flows.box_18  # type: ignore erroneous pylance subscript error
    reported_flows_sheet["I10"] = reported_flows.box_18  # type: ignore erroneous pylance subscript error

    # Roughbark Reservoir
    summary_sheet["G26"] = boxes["box19"]  # type: ignore erroneous pylance subscript error
    summary_sheet["H26"] = boxes["box20"]  # type: ignore erroneous pylance subscript error
    summary_sheet["I26"] = boxes["box21"]  # type: ignore erroneous pylance subscript error

    # Rafferty Reservoir
    summary_sheet["J26"] = boxes["box22"]  # type: ignore erroneous pylance subscript error
    summary_sheet["K24"] = boxes["box23a"]  # type: ignore erroneous pylance subscript error
    summary_sheet["K26"] = reported_flows.box_5b  # type: ignore erroneous pylance subscript error # Same as pipeline
    summary_sheet["L26"] = boxes["box24"]  # type: ignore erroneous pylance subscript error

    # Minor Project Diversions
    summary_sheet["M26"] = reported_flows.box_25  # type: ignore erroneous pylance subscript error
    reported_flows_sheet["J10"] = reported_flows.box_25  # type: ignore erroneous pylance subscript error

    # Total Diversions Upper Souris
    summary_sheet["N26"] = boxes["box26"]  # type: ignore erroneous pylance subscript error

    # Lower Souris
    summary_sheet["B37"] = reported_flows.box_27  # type: ignore erroneous pylance subscript error
    reported_flows_sheet["K10"] = reported_flows.box_27  # type: ignore erroneous pylance subscript error
    summary_sheet["C37"] = reported_flows.box_28  # type: ignore erroneous pylance subscript error
    reported_flows_sheet["L10"] = reported_flows.box_28  # type: ignore erroneous pylance subscript error
    summary_sheet["D37"] = reported_flows.box_29  # type: ignore erroneous pylance subscript error
    reported_flows_sheet["M10"] = reported_flows.box_29  # type: ignore erroneous pylance subscript error
    summary_sheet["E37"] = boxes["box30"]  # type: ignore erroneous pylance subscript error

    # Moose Mountain
    summary_sheet["G37"] = boxes["box31"]  # type: ignore erroneous pylance subscript error
    summary_sheet["H37"] = boxes["box32"]  # type: ignore erroneous pylance subscript error
    summary_sheet["I37"] = boxes["box33"]  # type: ignore erroneous pylance subscript error

    # Grant Divine
    summary_sheet["J37"] = boxes["box34"]  # type: ignore erroneous pylance subscript error
    summary_sheet["K37"] = boxes["box35"]  # type: ignore erroneous pylance subscript error
    summary_sheet["L37"] = boxes["box36"]  # type: ignore erroneous pylance subscript error

    # Minor Project Diversions and Total Diversions Moose Creek
    summary_sheet["M37"] = reported_flows.box_37  # type: ignore erroneous pylance subscript error
    reported_flows_sheet["N10"] = reported_flows.box_37  # type: ignore erroneous pylance subscript error
    summary_sheet["N37"] = boxes["box38"]  # type: ignore erroneous pylance subscript error

    # Non-Contributory Basins
    summary_sheet["B48"] = boxes["box39"]  # type: ignore erroneous pylance subscript error
    summary_sheet["C48"] = boxes["box40"]  # type: ignore erroneous pylance subscript error
    summary_sheet["D48"] = boxes["box41"]  # type: ignore erroneous pylance subscript error

    # Summary of Natural Flow
    summary_sheet["F50"] = boxes["box42"]  # type: ignore erroneous pylance subscript error
    summary_sheet["G50"] = boxes["box43"]  # type: ignore erroneous pylance subscript error
    # reported_flows_sheet["O10"] = boxes["box43"]  # type: ignore erroneous pylance subscript error
    summary_sheet["H50"] = boxes["box44"]  # type: ignore erroneous pylance subscript error
    summary_sheet["I49"] = boxes["box45a"]  # type: ignore erroneous pylance subscript error
    summary_sheet["I51"] = boxes["box45b"]  # type: ignore erroneous pylance subscript error
    summary_sheet["J50"] = boxes["box46"]  # type: ignore erroneous pylance subscript error
    summary_sheet["K49"] = boxes["box47a"]  # type: ignore erroneous pylance subscript error
    summary_sheet["K51"] = boxes["box47b"]  # type: ignore erroneous pylance subscript error

    # Recommendation Section 2
    summary_sheet["M48"] = boxes["box48"]  # type: ignore erroneous pylance subscript error
    summary_sheet["N48"] = boxes["box49"]  # type: ignore erroneous pylance subscript error
    # reported_flows_sheet["P10"] = boxes["box49"]  # type: ignore erroneous pylance subscript error
    summary_sheet["O48"] = boxes["box50"]  # type: ignore erroneous pylance subscript error

    # --------------------------------------------------------------------------------------------------#
    #                                       INI Data                                                    #
    # --------------------------------------------------------------------------------------------------#
    #  These cells are indexed starting at 1 and not 0.  Weird.
    for idx, setting in enumerate(reported_flows.__dict__.items(), start=2):
        config_sheet.cell(row=idx, column=1, value=setting[0])
        config_sheet.cell(row=idx, column=2, value=setting[1])

    newest_log = get_newest_log(Path("logs"))
    with open(newest_log, mode="r") as log_file:
        for idx, line in enumerate(log_file, start=2):
            log_sheet.cell(row=idx, column=1, value=line)

    # --------------------------------------------------------------------------------------------------#
    #                                       Approval levels                                             #
    # --------------------------------------------------------------------------------------------------#
    # Reservoirs
    ts_data_sheet["B4"].value = approval_dict.get("05NA006", "Unknown")
    ts_data_sheet["C4"].value = approval_dict.get("05NB020", "Unknown")
    ts_data_sheet["D4"].value = approval_dict.get("05NB016", "Unknown")
    ts_data_sheet["E4"].value = approval_dict.get("05NC002", "Unknown")
    ts_data_sheet["F4"].value = approval_dict.get("05ND012", "Unknown")

    # Discharge
    ts_data_sheet["M4"].value = approval_dict.get("05NB001", "Unknown")
    ts_data_sheet["N4"].value = approval_dict.get("05NB036", "Unknown")
    ts_data_sheet["O4"].value = approval_dict.get("05NB011", "Unknown")
    ts_data_sheet["P4"].value = approval_dict.get("05NB018", "Unknown")
    ts_data_sheet["Q4"].value = approval_dict.get("05NA003", "Unknown")
    ts_data_sheet["R4"].value = approval_dict.get("05NB040", "Unknown")
    ts_data_sheet["S4"].value = approval_dict.get("05NB041", "Unknown")
    ts_data_sheet["T4"].value = approval_dict.get("05NB038", "Unknown")
    ts_data_sheet["U4"].value = approval_dict.get("05NB014", "Unknown")
    ts_data_sheet["V4"].value = approval_dict.get("05NB035", "Unknown")
    ts_data_sheet["W4"].value = approval_dict.get("05NB033", "Unknown")
    ts_data_sheet["X4"].value = approval_dict.get("05NB039", "Unknown")
    ts_data_sheet["Y4"].value = approval_dict.get("05113600", "Unknown")
    ts_data_sheet["Z4"].value = approval_dict.get("05114000", "Unknown")

    # Roughbark
    ts_data_sheet["AB4"].value = approval_dict.get("05NB016_wind_speed", "Unknown")
    ts_data_sheet["AC4"].value = approval_dict.get("05NB016_radiation", "Unknown")
    ts_data_sheet["AD4"].value = approval_dict.get("05NB016_temperature", "Unknown")
    ts_data_sheet["AE4"].value = approval_dict.get("05NB016_rel_humidity", "Unknown")
    ts_data_sheet["AF4"].value = approval_dict.get("05NB016_precip", "Unknown")

    # Handsworth
    ts_data_sheet["AI4"].value = approval_dict.get("05NCM01_wind_speed", "Unknown")
    ts_data_sheet["AJ4"].value = approval_dict.get("05NCM01_radiation", "Unknown")
    ts_data_sheet["AK4"].value = approval_dict.get("05NCM01_temperature", "Unknown")
    ts_data_sheet["AL4"].value = approval_dict.get("05NCM01_rel_humidity", "Unknown")
    ts_data_sheet["AM4"].value = approval_dict.get("05NCM01_precip", "Unknown")
    wb.save(filename)

    # --------------------------------------------------------------------------------------------------#
    #                                       TS Data                                                     #
    # --------------------------------------------------------------------------------------------------#
    writer = pd.ExcelWriter(
        filename,
        engine="openpyxl",
        mode="a",
        if_sheet_exists="overlay",
        date_format="YYYY-MM-DD",
        datetime_format="YYYY-MM-DD",
    )

    writer.book = wb

    # Discharge data
    daily_data[
        [
            "05NB001_discharge",
            "05NB036_discharge",
            "05NB011_discharge",
            "05NB018_discharge",
            "05NA003_discharge",
            "05NB040_discharge",
            "05NB041_discharge",
            "05NB038_discharge",
            "05NB014_discharge",
            "05NB035_discharge",
            "05NB033_discharge",
            "05NB039_discharge",
            "05113600_discharge",
            "05114000_discharge",
        ]
    ].to_excel(
        writer,
        sheet_name=ts_data_sheet_name,
        startrow=5,
        startcol=12,
        header=False,
        index=False,
    )

    # Reservoir data
    daily_data[
        [
            "05NA006_elevation",
            "05NB020_elevation",
            "05NB016_elevation",
            "05NC002_elevation",
            "05ND012_elevation",
        ]
    ].to_excel(
        writer,
        sheet_name=ts_data_sheet_name,
        startrow=5,
        header=False,
        index=True,
    )

    monthly_elev_data["05NA006"].to_excel(
        writer,
        sheet_name=ts_data_sheet_name,
        startrow=6,
        startcol=9,
        header=False,
        index=False,
    )
    monthly_elev_data["05NB020"].to_excel(
        writer,
        sheet_name=ts_data_sheet_name,
        startrow=21,
        startcol=9,
        header=False,
        index=False,
    )
    monthly_elev_data["05NB016"].to_excel(
        writer,
        sheet_name=ts_data_sheet_name,
        startrow=36,
        startcol=9,
        header=False,
        index=False,
    )
    monthly_elev_data["05NC002"].to_excel(
        writer,
        sheet_name=ts_data_sheet_name,
        startrow=51,
        startcol=9,
        header=False,
        index=False,
    )
    monthly_elev_data["05ND012"].to_excel(
        writer,
        sheet_name=ts_data_sheet_name,
        startrow=66,
        startcol=9,
        header=False,
        index=False,
    )

    # Meteo data
    daily_meteo_data["05NB016"].to_excel(
        writer,
        sheet_name=ts_data_sheet_name,
        startrow=5,
        startcol=27,
        header=False,
        index=False,
    )
    daily_meteo_data["05NCM01"].to_excel(
        writer,
        sheet_name=ts_data_sheet_name,
        startrow=5,
        startcol=34,
        header=False,
        index=False,
    )
    daily_meteo_data["oxbow"][["oxbow_precip"]].to_excel(
        writer,
        sheet_name=ts_data_sheet_name,
        startrow=5,
        startcol=40,
        header=False,
        index=False,
    )

    if override_data is not None:
        override_data.to_excel(
            writer,
            sheet_name=original_ts_sheet_name,
            startrow=5,
            startcol=0,
        )

    # Format date.  Tried other methods, this is the one I found to work
    worksheet = writer.sheets[ts_data_sheet_name]
    for row in worksheet.iter_rows(min_row=4, max_row=400, max_col=1):
        for cell in row:
            cell.number_format = "yyyy-mm-dd"

    worksheet = writer.sheets[original_ts_sheet_name]
    for row in worksheet.iter_rows(min_row=7, max_row=400, max_col=1):
        for cell in row:
            cell.number_format = "yyyy-mm-dd"
    writer.save()
    return writer


def fill_gaps_nan(
    dataframe: pd.DataFrame,
    start_date: str,
    end_date: str,
    freq: str,
) -> pd.DataFrame:
    """Fill gaps in DateTimeIndex'ed DataFrame with NaN values.

    Parameters
    ----------
    dataframe : pd.DataFrame
        Time Series data with DateTimeIndex
    start_date : str
        Expected start date of data
    end_date : str
        Expected end date of data
    freq : str
        Frequency to check for data gaps

    Returns
    -------
    pd.DataFrame
        DataFrame with NaN values where missing data was detected.
    """
    dt_index = pd.date_range(start_date, end_date, freq=freq)
    index_name = dataframe.index.name
    dataframe = dataframe.reindex(dt_index).fillna(float("NaN"))
    dataframe.index.name = index_name
    return dataframe


def merge_override_data(original: dict, override: dict) -> dict:
    """Merge override data into original DataFrames

    Parameters
    ----------
    left_dict : dict
        Dictionary of original DataFrames
    right_dict : dict
        Dictionary of override DataFrames

    Returns
    -------
    dict
        Dictionary of original frames with override data applied if override
        data was provided in the right_dict.
    Notes
    -----
    Not the most clever function, but it appears to get the job done.
    """
    if not override:
        return original
    # Ignore any override data that doesn't have a corrosponding left_dict key.
    # override = {staid: data for staid, data in override.items() if staid in original}
    for staid, dataframe in override.items():
        # concat fills gaps with new values
        original[staid] = pd.concat([original[staid], dataframe[~dataframe.index.isin(original[staid].index)]])
        # update fills existing values
        original[staid].update(dataframe)
        original[staid] = original[staid].sort_index()
    return original


def debug_boxes_df(data: dict) -> pd.DataFrame:
    drop_list = (
        "box4",
        "box7",
        "box11",
        "box12",
        "box16",
        "box18",
        "box25",
        "box27",
        "box28",
        "box29",
        "box37",
        "box43",
    )
    data = {box: value for box, value in data.items() if box not in drop_list}
    df = pd.DataFrame(data.items(), columns=["box_num", "value_dam3"])
    df.replace({"box_num": "box47a"}, "box47", inplace=True)
    return df
