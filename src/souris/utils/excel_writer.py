from datetime import datetime as dt
from pathlib import Path
from typing import Union
from io import BytesIO
from openpyxl.utils.dataframe import dataframe_to_rows

import pandas as pd
from openpyxl import load_workbook

import src.souris.core.boxes as bx
import src.souris.core.dates as dates


def get_newest_log(path: Path) -> Path:
    """Finds most recent file in directory.

    Parameters
    ----------
    path : Path
        pathlib.Path filepath to directory to search

    Returns
    -------
    Path
        pathlib.Path of most recent file in directory
    """
    files = list(path.iterdir())
    return max(files, key=lambda file: file.stat().st_mtime)


def souris_excel_writer(
    dates: dates.Dates,
    boxes: bx.Boxes,
    daily_discharge: pd.DataFrame,
    daily_reservoir: dict[str : pd.DataFrame],
    # monthly_reservoir: pd.DataFrame,
    daily_roughbark: pd.DataFrame,
    daily_handsworth: pd.DataFrame,
    daily_oxbow: pd.DataFrame,
    # override_data: Union[None, pd.DataFrame],
):

    excel_stream = BytesIO()
    report_template = Path("souris/data/xlsx_template/BLANK_souris_natural_flow_apportionment_report.xlsx")

    discharge_sheet_name = "Discharge Data"
    reservoir_sheet_name = "Reservoir Data"
    met_sheet_name = "Meteo Data"
    monthyl_sheet_name = "Monthly Tables"
    raw_sheet_name = "Raw Data"

    wb = load_workbook(report_template)
    summary_sheet = wb["Souris Natural Flow Summary"]
    reported_flows_sheet = wb["Input Reported Flows"]
    log_sheet = wb["Log"]
    discharge_sheet = wb[discharge_sheet_name]
    reservoir_sheet = wb[reservoir_sheet_name]
    raw_data_sheet = wb[raw_sheet_name]

    # Header
    app_range = f"for {dates.start_apportion} to {dates.end_apportion}"
    summary_sheet["B5"] = f"For the period of {app_range}"  # type: ignore erroneous pylance subscript error

    # --------------------------------------------------------------------------------------------------#
    #                                       Boxes                                                       #
    # --------------------------------------------------------------------------------------------------#

    # Larson Reservoir
    summary_sheet["B15"] = boxes.box_1  # type: ignore erroneous pylance subscript error
    summary_sheet["C15"] = boxes.box_2  # type: ignore erroneous pylance subscript error
    summary_sheet["D15"] = boxes.box_3  # type: ignore erroneous pylance subscript error
    # summary_sheet["E15"] = boxes["box4"]  # type: ignore erroneous pylance subscript error
    # reported_flows_sheet["C10"] = boxes["box4"]  # type: ignore erroneous pylance subscript error

    # Boundary Reservoir
    summary_sheet["F14"] = boxes.box_5a  # type: ignore erroneous pylance subscript error
    summary_sheet["F15"] = 0  # type: ignore erroneous pylance subscript error
    # reported_flows_sheet["D10"] = boxes["box5a"]  # type: ignore erroneous pylance subscript error
    reported_flows_sheet["D10"] = 0  # type: ignore erroneous pylance subscript error
    summary_sheet["G15"] = boxes.box_6  # type: ignore erroneous pylance subscript error

    # summary_sheet["H15"] = boxes["box7"]  # type: ignore erroneous pylance subscript error
    # reported_flows_sheet["E10"] = boxes["box7"]  # type: ignore erroneous pylance subscript error
    summary_sheet["I15"] = boxes.box_8  # type: ignore erroneous pylance subscript error
    summary_sheet["J15"] = boxes.box_9  # type: ignore erroneous pylance subscript error
    summary_sheet["K15"] = boxes.box_10  # type: ignore erroneous pylance subscript error
    summary_sheet["L15"] = boxes.box_11  # type: ignore erroneous pylance subscript error
    reported_flows_sheet["F10"] = boxes.box_11  # type: ignore erroneous pylance subscript error
    summary_sheet["M15"] = boxes.box_12  # type: ignore erroneous pylance subscript error
    reported_flows_sheet["G10"] = boxes.box_12  # type: ignore erroneous pylance subscript error
    summary_sheet["N15"] = boxes.box_13  # type: ignore erroneous pylance subscript error

    # Nickle Lake Reservoir
    summary_sheet["B26"] = boxes.box_14  # type: ignore erroneous pylance subscript error
    summary_sheet["C26"] = boxes.box_15  # type: ignore erroneous pylance subscript error
    summary_sheet["D26"] = boxes.box_16  # type: ignore erroneous pylance subscript error
    reported_flows_sheet["H10"] = boxes.box_16  # type: ignore erroneous pylance subscript error
    summary_sheet["E26"] = boxes.box_17  # type: ignore erroneous pylance subscript error

    # City of Wayburn Return Flow
    summary_sheet["F26"] = boxes.box_18  # type: ignore erroneous pylance subscript error
    reported_flows_sheet["I10"] = boxes.box_18  # type: ignore erroneous pylance subscript error

    # Roughbark Reservoir
    summary_sheet["G26"] = boxes.box_19  # type: ignore erroneous pylance subscript error
    summary_sheet["H26"] = boxes.box_20  # type: ignore erroneous pylance subscript error
    summary_sheet["I26"] = boxes.box_21  # type: ignore erroneous pylance subscript error

    # Rafferty Reservoir
    summary_sheet["J26"] = boxes.box_22  # type: ignore erroneous pylance subscript error
    summary_sheet["K24"] = boxes.box_23a  # type: ignore erroneous pylance subscript error
    summary_sheet["K26"] = boxes.box_5b  # type: ignore erroneous pylance subscript error # Same as pipeline
    summary_sheet["L26"] = boxes.box_24  # type: ignore erroneous pylance subscript error

    # Minor Project Diversions
    summary_sheet["M26"] = boxes.box_25  # type: ignore erroneous pylance subscript error
    reported_flows_sheet["J10"] = boxes.box_25  # type: ignore erroneous pylance subscript error

    # Total Diversions Upper Souris
    summary_sheet["N26"] = boxes.box_26  # type: ignore erroneous pylance subscript error

    # Lower Souris
    summary_sheet["B37"] = boxes.box_27  # type: ignore erroneous pylance subscript error
    reported_flows_sheet["K10"] = boxes.box_27  # type: ignore erroneous pylance subscript error
    summary_sheet["C37"] = boxes.box_28  # type: ignore erroneous pylance subscript error
    reported_flows_sheet["L10"] = boxes.box_28  # type: ignore erroneous pylance subscript error
    summary_sheet["D37"] = boxes.box_29  # type: ignore erroneous pylance subscript error
    reported_flows_sheet["M10"] = boxes.box_29  # type: ignore erroneous pylance subscript error
    summary_sheet["E37"] = boxes.box_30  # type: ignore erroneous pylance subscript error

    # Moose Mountain
    summary_sheet["G37"] = boxes.box_31  # type: ignore erroneous pylance subscript error
    summary_sheet["H37"] = boxes.box_32  # type: ignore erroneous pylance subscript error
    summary_sheet["I37"] = boxes.box_33  # type: ignore erroneous pylance subscript error

    # Grant Divine
    summary_sheet["J37"] = boxes.box_34  # type: ignore erroneous pylance subscript error
    summary_sheet["K37"] = boxes.box_35  # type: ignore erroneous pylance subscript error
    summary_sheet["L37"] = boxes.box_36  # type: ignore erroneous pylance subscript error

    # Minor Project Diversions and Total Diversions Moose Creek
    summary_sheet["M37"] = boxes.box_37  # type: ignore erroneous pylance subscript error
    reported_flows_sheet["N10"] = boxes.box_37  # type: ignore erroneous pylance subscript error
    summary_sheet["N37"] = boxes.box_38  # type: ignore erroneous pylance subscript error

    # Non-Contributory Basins
    summary_sheet["B48"] = boxes.box_39  # type: ignore erroneous pylance subscript error
    summary_sheet["C48"] = boxes.box_40  # type: ignore erroneous pylance subscript error
    summary_sheet["D48"] = boxes.box_41  # type: ignore erroneous pylance subscript error

    # Summary of Natural Flow
    summary_sheet["F50"] = boxes.box_42  # type: ignore erroneous pylance subscript error
    summary_sheet["G50"] = boxes.box_43  # type: ignore erroneous pylance subscript error
    # reported_flows_sheet["O10"] = boxes["box43"]  # type: ignore erroneous pylance subscript error
    summary_sheet["H50"] = boxes.box_44  # type: ignore erroneous pylance subscript error
    summary_sheet["I49"] = boxes.box_45a  # type: ignore erroneous pylance subscript error
    summary_sheet["I51"] = boxes.box_45b  # type: ignore erroneous pylance subscript error
    summary_sheet["J50"] = boxes.box_46  # type: ignore erroneous pylance subscript error
    summary_sheet["K49"] = boxes.box_47a  # type: ignore erroneous pylance subscript error
    summary_sheet["K51"] = boxes.box_47b  # type: ignore erroneous pylance subscript error

    # Recommendation Section 2
    summary_sheet["M48"] = boxes.box_48  # type: ignore erroneous pylance subscript error
    summary_sheet["N48"] = boxes.box_49  # type: ignore erroneous pylance subscript error
    # reported_flows_sheet["P10"] = boxes["box49"]  # type: ignore erroneous pylance subscript error
    summary_sheet["O48"] = boxes.box_50  # type: ignore erroneous pylance subscript error

    # --------------------------------------------------------------------------------------------------#
    #                                       INI Data                                                    #
    # --------------------------------------------------------------------------------------------------#
    #  These cells are indexed starting at 1 and not 0.  Weird.
    # for idx, setting in enumerate(reported_flows.__dict__.items(), start=2):
    #     config_sheet.cell(row=idx, column=1, value=setting[0])
    #     config_sheet.cell(row=idx, column=2, value=setting[1])

    # newest_log = get_newest_log(Path("logs"))
    # with open(newest_log, mode="r") as log_file:
    #     for idx, line in enumerate(log_file, start=2):
    #         log_sheet.cell(row=idx, column=1, value=line)
    # wb.save(filename)

    """Need to figure out how to write pandas dataframes to excel while still open with pyxl.  look at dateframe_to_rows util"""
    # --------------------------------------------------------------------------------------------------#
    #                                       TS Data                                                     #
    # --------------------------------------------------------------------------------------------------#
    # Discharge data
    daily_discharge.index = daily_discharge.index.astype(str)
    discharge_rows = dataframe_to_rows(daily_discharge, index=True, header=False)
    for r_idx, row in enumerate(discharge_rows, 3):
        for c_idx, value in enumerate(row, 1):
            discharge_sheet.cell(row=r_idx, column=c_idx, value=value)

    # # Reservoir data
    daily_reservoir["05NA006"].index = daily_reservoir["05NA006"].index.astype(str)
    reservoir_rows = dataframe_to_rows(daily_reservoir["05NA006"], index=True, header=False)
    for r_idx, row in enumerate(reservoir_rows, 3):
        for c_idx, value in enumerate(row, 1):
            reservoir_sheet.cell(row=r_idx, column=c_idx, value=value)

    daily_reservoir["05NB020"].index = daily_reservoir["05NB020"].index.astype(str)
    reservoir_rows = dataframe_to_rows(daily_reservoir["05NB020"], index=False, header=False)
    for r_idx, row in enumerate(reservoir_rows, 4):
        for c_idx, value in enumerate(row, 6):
            reservoir_sheet.cell(row=r_idx, column=c_idx, value=value)

    daily_reservoir["05NB016"].index = daily_reservoir["05NB016"].index.astype(str)
    reservoir_rows = dataframe_to_rows(daily_reservoir["05NB016"], index=False, header=False)
    for r_idx, row in enumerate(reservoir_rows, 4):
        for c_idx, value in enumerate(row, 10):
            reservoir_sheet.cell(row=r_idx, column=c_idx, value=value)

    daily_reservoir["05NC002"].index = daily_reservoir["05NC002"].index.astype(str)
    reservoir_rows = dataframe_to_rows(daily_reservoir["05NC002"], index=False, header=False)
    for r_idx, row in enumerate(reservoir_rows, 4):
        for c_idx, value in enumerate(row, 14):
            reservoir_sheet.cell(row=r_idx, column=c_idx, value=value)

    daily_reservoir["05ND012"].index = daily_reservoir["05ND012"].index.astype(str)
    reservoir_rows = dataframe_to_rows(daily_reservoir["05ND012"], index=False, header=False)
    for r_idx, row in enumerate(reservoir_rows, 4):
        for c_idx, value in enumerate(row, 18):
            reservoir_sheet.cell(row=r_idx, column=c_idx, value=value)

    # daily_reservoir.to_excel(
    #     writer,
    #     sheet_name=discharge_sheet_name,
    #     startrow=5,
    #     header=False,
    #     index=True,
    # )

    # monthly_elev_data["05NA006"].to_excel(
    #     writer,
    #     sheet_name=discharge_sheet_name,
    #     startrow=6,
    #     startcol=9,
    #     header=False,
    #     index=False,
    # )
    # monthly_elev_data["05NB020"].to_excel(
    #     writer,
    #     sheet_name=discharge_sheet_name,
    #     startrow=21,
    #     startcol=9,
    #     header=False,
    #     index=False,
    # )
    # monthly_elev_data["05NB016"].to_excel(
    #     writer,
    #     sheet_name=discharge_sheet_name,
    #     startrow=36,
    #     startcol=9,
    #     header=False,
    #     index=False,
    # )
    # monthly_elev_data["05NC002"].to_excel(
    #     writer,
    #     sheet_name=discharge_sheet_name,
    #     startrow=51,
    #     startcol=9,
    #     header=False,
    #     index=False,
    # )
    # monthly_elev_data["05ND012"].to_excel(
    #     writer,
    #     sheet_name=discharge_sheet_name,
    #     startrow=66,
    #     startcol=9,
    #     header=False,
    #     index=False,
    # )

    # Meteo data
    # daily_roughbark["05NB016"].to_excel(
    #     writer,
    #     sheet_name=discharge_sheet_name,
    #     startrow=5,
    #     startcol=27,
    #     header=False,
    #     index=False,
    # )
    # daily_roughbark["05NCM01"].to_excel(
    #     writer,
    #     sheet_name=discharge_sheet_name,
    #     startrow=5,
    #     startcol=34,
    #     header=False,
    #     index=False,
    # )
    # daily_roughbark["oxbow"][["oxbow_precip"]].to_excel(
    #     writer,
    #     sheet_name=discharge_sheet_name,
    #     startrow=5,
    #     startcol=40,
    #     header=False,
    #     index=False,
    # )

    # if override_data is not None:
    #     override_data.to_excel(
    #         writer,
    #         sheet_name=raw_sheet_name,
    #         startrow=5,
    #         startcol=0,
    #     )

    # # Format date.  Tried other methods, this is the one I found to work
    # worksheet = writer.sheets[discharge_sheet_name]
    # for row in worksheet.iter_rows(min_row=4, max_row=400, max_col=1):
    #     for cell in row:
    #         cell.number_format = "yyyy-mm-dd"

    # worksheet = writer.sheets[raw_sheet_name]
    # for row in worksheet.iter_rows(min_row=7, max_row=400, max_col=1):
    #     for cell in row:
    #         cell.number_format = "yyyy-mm-dd"
    wb.save(excel_stream)
    excel_stream.seek(0)
    return excel_stream


def fill_gaps_nan(
    dataframe: pd.DataFrame,
    start_date: str,
    end_date: str,
    freq: str,
) -> pd.DataFrame:
    """Fill gaps in DateTimeIndex'ed DataFrame with NaN values.

    Parameters
    ----------
    dataframe : pd.DataFrame
        Time Series data with DateTimeIndex
    start_date : str
        Expected start date of data
    end_date : str
        Expected end date of data
    freq : str
        Frequency to check for data gaps

    Returns
    -------
    pd.DataFrame
        DataFrame with NaN values where missing data was detected.
    """
    dt_index = pd.date_range(start_date, end_date, freq=freq)
    index_name = dataframe.index.name
    dataframe = dataframe.reindex(dt_index).fillna(float("NaN"))
    dataframe.index.name = index_name
    return dataframe


def merge_override_data(original: dict, override: dict) -> dict:
    """Merge override data into original DataFrames

    Parameters
    ----------
    left_dict : dict
        Dictionary of original DataFrames
    right_dict : dict
        Dictionary of override DataFrames

    Returns
    -------
    dict
        Dictionary of original frames with override data applied if override
        data was provided in the right_dict.
    Notes
    -----
    Not the most clever function, but it appears to get the job done.
    """
    if not override:
        return original
    # Ignore any override data that doesn't have a corrosponding left_dict key.
    # override = {staid: data for staid, data in override.items() if staid in original}
    for staid, dataframe in override.items():
        # concat fills gaps with new values
        original[staid] = pd.concat([original[staid], dataframe[~dataframe.index.isin(original[staid].index)]])
        # update fills existing values
        original[staid].update(dataframe)
        original[staid] = original[staid].sort_index()
    return original
