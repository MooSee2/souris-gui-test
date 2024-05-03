import os
from datetime import datetime as dt
from io import BytesIO
from pathlib import Path
from typing import Union

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

import src.souris_model.core.boxes as bx
import src.souris_model.core.dates as dates


def get_newest_log(path: Path) -> Path:
    """Finds most recent file in directory.

    Parameters
    ----------
    path : Path
        pathlib.Path filepath to directory to search

    Returns
    -------
    Path
        pathlib.Path of most recent file in directory
    """
    files = list(path.iterdir())
    return max(files, key=lambda file: file.stat().st_mtime)


def df_writer(
    df: pd.DataFrame,
    sheet,
    start_row: int = 0,
    start_col: int = 0,
    index=False,
    header=False,
):
    df.index = df.index.astype(str)
    rows = dataframe_to_rows(df, index=index, header=header)
    for r_idx, row in enumerate(rows, start_row):
        for c_idx, value in enumerate(row, start_col):
            sheet.cell(row=r_idx, column=c_idx, value=value)
    return None


def souris_excel_writer(
    dates: dates.Dates,
    boxes: bx.Boxes,
    daily_discharge: pd.DataFrame,
    raw_discharge_reservoirs: pd.DataFrame,
    daily_reservoir: dict[str : pd.DataFrame],
    daily_roughbark: pd.DataFrame,
    daily_handsworth: pd.DataFrame,
    daily_oxbow: pd.DataFrame,
    monthly_reservoir_SAC: dict[str : pd.DataFrame],
    monthly_roughbark_evap_precip: pd.DataFrame,
    monthly_handsworth_evap_precip: pd.DataFrame,
    monthly_oxbow_precip: pd.DataFrame,
    reservoir_losses: dict[str : pd.DataFrame],
    report_template=Path("souris_model/data/xlsx_template/BLANK_souris_natural_flow_apportionment_report.xlsx"),
):

    # debug_excel = os.getenv("DEBUG_EXCEL", False).lower() in ("true", "1", "t")
    debug_excel = False
    excel_stream = BytesIO()

    discharge_sheet_name = "Discharge Data"
    reservoir_sheet_name = "Reservoir Data"
    met_sheet_name = "Meteo Data"
    monthly_sheet_name = "Monthly Tables"
    raw_sheet_name = "Raw Data"

    wb = load_workbook(report_template)
    summary_sheet = wb["Souris Natural Flow Summary"]
    reported_flows_sheet = wb["Input Reported Flows"]
    log_sheet = wb["Log"]
    discharge_sheet = wb[discharge_sheet_name]
    reservoir_sheet = wb[reservoir_sheet_name]
    met_sheet = wb[met_sheet_name]
    monthly_tables_sheet = wb[monthly_sheet_name]
    raw_data_sheet = wb[raw_sheet_name]

    # Header
    app_range = f"for {dates.start_apportion} to {dates.end_apportion}"
    summary_sheet["B5"] = f"For the period of {app_range}"

    # --------------------------------------------------------------------------------------------------#
    #                                       Boxes                                                       #
    # --------------------------------------------------------------------------------------------------#

    # Larson Reservoir
    summary_sheet["B15"] = boxes.box_1
    summary_sheet["C15"] = boxes.box_2
    summary_sheet["D15"] = boxes.box_3
    # summary_sheet["E15"] = boxes["box4"]
    # reported_flows_sheet["C10"] = boxes["box4"]

    # Boundary Reservoir
    summary_sheet["F14"] = boxes.box_5a
    summary_sheet["F15"] = 0
    # reported_flows_sheet["D10"] = boxes["box5a"]
    reported_flows_sheet["D10"] = 0
    summary_sheet["G15"] = boxes.box_6

    # summary_sheet["H15"] = boxes["box7"]
    # reported_flows_sheet["E10"] = boxes["box7"]
    summary_sheet["I15"] = boxes.box_8
    summary_sheet["J15"] = boxes.box_9
    summary_sheet["K15"] = boxes.box_10
    summary_sheet["L15"] = boxes.box_11
    reported_flows_sheet["F10"] = boxes.box_11
    summary_sheet["M15"] = boxes.box_12
    reported_flows_sheet["G10"] = boxes.box_12
    summary_sheet["N15"] = boxes.box_13

    # Nickle Lake Reservoir
    summary_sheet["B26"] = boxes.box_14
    summary_sheet["C26"] = boxes.box_15
    summary_sheet["D26"] = boxes.box_16
    reported_flows_sheet["H10"] = boxes.box_16
    summary_sheet["E26"] = boxes.box_17

    # City of Wayburn Return Flow
    summary_sheet["F26"] = boxes.box_18
    reported_flows_sheet["I10"] = boxes.box_18

    # Roughbark Reservoir
    summary_sheet["G26"] = boxes.box_19
    summary_sheet["H26"] = boxes.box_20
    summary_sheet["I26"] = boxes.box_21

    # Rafferty Reservoir
    summary_sheet["J26"] = boxes.box_22
    summary_sheet["K24"] = boxes.box_23a
    summary_sheet["K26"] = boxes.box_5b
    summary_sheet["L26"] = boxes.box_24

    # Minor Project Diversions
    summary_sheet["M26"] = boxes.box_25
    reported_flows_sheet["J10"] = boxes.box_25

    # Total Diversions Upper Souris
    summary_sheet["N26"] = boxes.box_26

    # Lower Souris
    summary_sheet["B37"] = boxes.box_27
    reported_flows_sheet["K10"] = boxes.box_27
    summary_sheet["C37"] = boxes.box_28
    reported_flows_sheet["L10"] = boxes.box_28
    summary_sheet["D37"] = boxes.box_29
    reported_flows_sheet["M10"] = boxes.box_29
    summary_sheet["E37"] = boxes.box_30

    # Moose Mountain
    summary_sheet["G37"] = boxes.box_31
    summary_sheet["H37"] = boxes.box_32
    summary_sheet["I37"] = boxes.box_33

    # Grant Divine
    summary_sheet["J37"] = boxes.box_34
    summary_sheet["K37"] = boxes.box_35
    summary_sheet["L37"] = boxes.box_36

    # Minor Project Diversions and Total Diversions Moose Creek
    summary_sheet["M37"] = boxes.box_37
    reported_flows_sheet["N10"] = boxes.box_37
    summary_sheet["N37"] = boxes.box_38

    # Non-Contributory Basins
    summary_sheet["B48"] = boxes.box_39
    summary_sheet["C48"] = boxes.box_40
    summary_sheet["D48"] = boxes.box_41

    # Summary of Natural Flow
    summary_sheet["F50"] = boxes.box_42
    summary_sheet["G50"] = boxes.box_43
    # reported_flows_sheet["O10"] = boxes["box43"]
    summary_sheet["H50"] = boxes.box_44
    summary_sheet["I49"] = boxes.box_45a
    summary_sheet["I51"] = boxes.box_45b
    summary_sheet["J50"] = boxes.box_46
    summary_sheet["K49"] = boxes.box_47a
    summary_sheet["K51"] = boxes.box_47b

    # Recommendation Section 2
    summary_sheet["M48"] = boxes.box_48
    summary_sheet["N48"] = boxes.box_49
    # reported_flows_sheet["P10"] = boxes["box49"]
    summary_sheet["O48"] = boxes.box_50

    # --------------------------------------------------------------------------------------------------#
    #                                       TS Data                                                     #
    # --------------------------------------------------------------------------------------------------#
    # Discharge data
    df_writer(daily_discharge, sheet=discharge_sheet, start_row=3, start_col=1, index=True)

    # # Reservoir data
    df_writer(daily_reservoir["05NA006"], sheet=reservoir_sheet, start_row=3, start_col=1, index=True)
    df_writer(daily_reservoir["05NB020"], sheet=reservoir_sheet, start_row=4, start_col=6, header=debug_excel)
    df_writer(daily_reservoir["05NB016"], sheet=reservoir_sheet, start_row=4, start_col=10, header=debug_excel)
    df_writer(daily_reservoir["05NC002"], sheet=reservoir_sheet, start_row=4, start_col=14, header=debug_excel)
    df_writer(daily_reservoir["05ND012"], sheet=reservoir_sheet, start_row=4, start_col=18, header=debug_excel)

    # Met data
    df_writer(daily_roughbark, sheet=met_sheet, start_row=3, start_col=1, index=True, header=debug_excel)
    df_writer(daily_handsworth, sheet=met_sheet, start_row=4, start_col=9, header=debug_excel)
    df_writer(daily_oxbow, sheet=met_sheet, start_row=4, start_col=16, header=debug_excel)

    # Monthly reservoir SAC
    df_writer(monthly_reservoir_SAC["05NA006"], sheet=monthly_tables_sheet, start_row=6, start_col=3)
    df_writer(monthly_reservoir_SAC["05NB020"], sheet=monthly_tables_sheet, start_row=21, start_col=3)
    df_writer(monthly_reservoir_SAC["05NB016"], sheet=monthly_tables_sheet, start_row=36, start_col=3)
    df_writer(monthly_reservoir_SAC["05NC002"], sheet=monthly_tables_sheet, start_row=51, start_col=3)
    df_writer(monthly_reservoir_SAC["05ND012"], sheet=monthly_tables_sheet, start_row=66, start_col=3)

    # Monthly reservoir loss
    df_writer(pd.DataFrame(reservoir_losses["05NA006"]), sheet=monthly_tables_sheet, start_row=6, start_col=9)
    df_writer(pd.DataFrame(reservoir_losses["05NB020"]), sheet=monthly_tables_sheet, start_row=21, start_col=9)
    df_writer(pd.DataFrame(reservoir_losses["05NB016"]), sheet=monthly_tables_sheet, start_row=36, start_col=9)
    df_writer(pd.DataFrame(reservoir_losses["05NC002"]), sheet=monthly_tables_sheet, start_row=51, start_col=9)
    df_writer(pd.DataFrame(reservoir_losses["05ND012"]), sheet=monthly_tables_sheet, start_row=66, start_col=9)

    # Roughbark, Handsworth precip/evap and oxbow precip
    df_writer(monthly_roughbark_evap_precip, sheet=monthly_tables_sheet, start_row=6, start_col=14)
    df_writer(monthly_handsworth_evap_precip, sheet=monthly_tables_sheet, start_row=21, start_col=14)
    df_writer(monthly_oxbow_precip, sheet=monthly_tables_sheet, start_row=35, start_col=14)

    # Raw data
    df_writer(raw_discharge_reservoirs, sheet=raw_data_sheet, start_row=1, start_col=1, index=True, header=True)

    # # Format date.  Tried other methods, this is the one I found to work
    # worksheet = writer.sheets[discharge_sheet_name]
    # for row in worksheet.iter_rows(min_row=4, max_row=400, max_col=1):
    #     for cell in row:
    #         cell.number_format = "yyyy-mm-dd"

    # worksheet = writer.sheets[raw_sheet_name]
    # for row in worksheet.iter_rows(min_row=7, max_row=400, max_col=1):
    #     for cell in row:
    #         cell.number_format = "yyyy-mm-dd"
    wb.save(excel_stream)
    excel_stream.seek(0)
    return excel_stream


def fill_gaps_nan(
    dataframe: pd.DataFrame,
    start_date: str,
    end_date: str,
    freq: str,
) -> pd.DataFrame:
    """Fill gaps in DateTimeIndex'ed DataFrame with NaN values.

    Parameters
    ----------
    dataframe : pd.DataFrame
        Time Series data with DateTimeIndex
    start_date : str
        Expected start date of data
    end_date : str
        Expected end date of data
    freq : str
        Frequency to check for data gaps

    Returns
    -------
    pd.DataFrame
        DataFrame with NaN values where missing data was detected.
    """
    dt_index = pd.date_range(start_date, end_date, freq=freq)
    index_name = dataframe.index.name
    dataframe = dataframe.reindex(dt_index).fillna(float("NaN"))
    dataframe.index.name = index_name
    return dataframe


def merge_override_data(original: dict, override: dict) -> dict:
    """Merge override data into original DataFrames

    Parameters
    ----------
    left_dict : dict
        Dictionary of original DataFrames
    right_dict : dict
        Dictionary of override DataFrames

    Returns
    -------
    dict
        Dictionary of original frames with override data applied if override
        data was provided in the right_dict.
    Notes
    -----
    Not the most clever function, but it appears to get the job done.
    """
    if not override:
        return original
    # Ignore any override data that doesn't have a corrosponding left_dict key.
    # override = {staid: data for staid, data in override.items() if staid in original}
    for staid, dataframe in override.items():
        # concat fills gaps with new values
        original[staid] = pd.concat([original[staid], dataframe[~dataframe.index.isin(original[staid].index)]])
        # update fills existing values
        original[staid].update(dataframe)
        original[staid] = original[staid].sort_index()
    return original
